"""
Excel 读取层：处理合并单元格、清洗数据类型

核心功能：
  - read_sheet(): 读取指定 sheet，返回清洗后的二维列表
  - read_sheet_as_dicts(): 以首行为列名返回 dict 列表
  - get_all_sheets(): 一次性读取所有 sheet
  - 自动处理合并单元格（填充 None）
  - 将 #DIV/0!、空字符串等转为安全值
"""

import openpyxl
from typing import Optional


def clean_value(value) -> Optional[float]:
    """
    清洗单元格值，将错误值、空字符串转为安全类型。

    Args:
        value: 原始单元格值

    Returns:
        清洗后的数值（float/int）或原始字符串，无法转换时返回 None
    """
    if value is None:
        return None

    # 处理字符串
    if isinstance(value, str):
        value = value.strip()
        if value in ("", "-", "#DIV/0!", "#N/A", "#VALUE!", "#REF!", "#NAME?", "#NUM!"):
            return None
        # 尝试转为数字
        try:
            return float(value)
        except ValueError:
            return value

    # 数值直接返回
    if isinstance(value, (int, float)):
        return value

    return None


def resolve_merged_cells(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    """
    构建合并单元格的填充映射字典。
    合并区域中非左上角的单元格 openpyxl 返回 None，
    此函数返回 {(row, col): value} 用于填充空单元格。

    Args:
        ws: openpyxl worksheet 对象

    Returns:
        {(row, col): value} 的字典
    """
    fill_map = {}
    for merge_range in ws.merged_cells.ranges:
        top_left_val = ws.cell(merge_range.min_row, merge_range.min_col).value
        if top_left_val is not None:
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    if row != merge_range.min_row or col != merge_range.min_col:
                        fill_map[(row, col)] = top_left_val
    return fill_map


def read_sheet(filepath: str, sheet_name: str) -> list[list]:
    """
    读取指定工作表，返回清洗后的二维列表。

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名称

    Returns:
        二维列表 data[row][col]，所有值已清洗
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"工作表 '{sheet_name}' 不存在。"
            f"可用工作表: {available}"
        )

    ws = wb[sheet_name]

    # 构建合并单元格填充映射
    fill_map = resolve_merged_cells(ws)

    # 读取数据
    result = []
    for row_idx in range(1, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            # 优先使用合并单元格填充值
            if (row_idx, col_idx) in fill_map:
                val = fill_map[(row_idx, col_idx)]
            else:
                val = ws.cell(row_idx, col_idx).value
            row_data.append(clean_value(val))
        result.append(row_data)

    wb.close()
    return result


def read_sheet_as_dicts(filepath: str, sheet_name: str,
                        header_row: int = 0) -> list[dict]:
    """
    以指定行为列名，返回 dict 列表。

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名称
        header_row: 表头行号（0-based，即 Excel 第1行为0）

    Returns:
        [{"列名": 值, ...}, ...]
    """
    data = read_sheet(filepath, sheet_name)
    if header_row >= len(data):
        return []

    headers = []
    for val in data[header_row]:
        headers.append(str(val) if val is not None else f"col_{len(headers)}")

    result = []
    for row in data[header_row + 1:]:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        # 跳过全空行
        if any(v is not None for v in record.values()):
            result.append(record)

    return result


def get_all_sheets(filepath: str) -> dict[str, list[list]]:
    """
    一次性读取所有工作表的清洗数据。

    Args:
        filepath: Excel 文件路径

    Returns:
        {sheet_name: [[...], ...]}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fill_map = resolve_merged_cells(ws)

        sheet_data = []
        for row_idx in range(1, ws.max_row + 1):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                if (row_idx, col_idx) in fill_map:
                    val = fill_map[(row_idx, col_idx)]
                else:
                    val = ws.cell(row_idx, col_idx).value
                row_data.append(clean_value(val))
            sheet_data.append(row_data)

        result[sheet_name] = sheet_data

    wb.close()
    return result


def safe_float(value, default: float = 0.0) -> float:
    """安全地将值转为 float，失败时返回 default。"""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default
